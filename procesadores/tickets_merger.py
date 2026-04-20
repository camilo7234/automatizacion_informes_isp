# procesadores/tickets_merger.py
# ------------------------------------------------------------------
# Genera el Excel contractual de PQRS (Calidad del Servicio).
# Formato: INFORMACION-A-REGISTRAR-POR-EL-ISP — hoja "PQRS"
# Estructura: filas = ID CUENTA | columnas = MES 1..24 × 5 subcategorías
# Fuente: CSV exportado desde Wispro (mesa de ayuda)
# ------------------------------------------------------------------

import json
import logging
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ==================================================================
# BLOQUE 1: CONFIGURACIÓN
# ==================================================================
RUTA_ENTRADA_WISPRO = Path("datos/entrada/wispro")          # ← cambia
RUTA_REGISTRO       = Path("datos/procesados/modelo_contrato/registro_procesados.json")
RUTA_SALIDA         = Path("salidas/tickets")

# Fecha de inicio del contrato → MES 1
CONTRATO_INICIO  = datetime(2025, 12, 1)
CONTRATO_MESES   = 24

# Mapeo Wispro Categoria → columna contractual
# Ajustar según los valores reales del CSV
MAPA_CATEGORIA = {
    # Disponibilidad
    "sin servicio":            "disponibilidad",
    "servicio intermitente":   "disponibilidad",
    "disponibilidad":          "disponibilidad",
    "caída de servicio":       "disponibilidad",
    # Velocidad
    "lentitud":                "velocidad",
    "velocidad":               "velocidad",
    "velocidad baja":          "velocidad",
    "baja velocidad":          "velocidad",
    # Falla red / CPE
    "falla en red":            "falla_cpe",
    "falla cpe":               "falla_cpe",
    "falla en red interna":    "falla_cpe",
    "falla equipo":            "falla_cpe",
    "soporte técnico":         "falla_cpe",
    # Traslado
    "traslado":                "traslado",
    "solicitud de traslado":   "traslado",
    "cambio de dirección":     "traslado",
}

SUBCOLS = ["pqr", "disponibilidad", "velocidad", "falla_cpe", "traslado"]

SUBCOLS_LABELS = {
    "pqr":            "PQR",
    "disponibilidad": "Disponibilidad\n(Sin servicio o\nservicio intermitente)",
    "velocidad":      "Velocidad de navegación\n(Muy baja con respecto\na lo comprometido)",
    "falla_cpe":      "Falla en red\ninterna o CPE",
    "traslado":       "Solicitud de\ntraslado",
}

ENCABEZADO_INFO = {
    "razon_social":  "EMPRESA MUNICIPAL DE TELECOMUNICACIONES DE IPIALES UNIMOS S.A E.S.P",
    "municipio":     "IPIALES",
    "departamento":  "NARIÑO",
    "contrato":      "AE2_40_2025",
    "fecha":         "5 DE DICIEMBRE DEL 2025",
}


class TicketsMerger:

    def __init__(
        self,
        ruta_entrada:  Path = RUTA_ENTRADA_WISPRO,          # ← cambia
        ruta_registro: Path = RUTA_REGISTRO,
        ruta_salida:   Path = RUTA_SALIDA,
    ):
        self.ruta_entrada  = Path(ruta_entrada)              # ← cambia
        self.ruta_registro = Path(ruta_registro)
        self.ruta_salida   = Path(ruta_salida)
        self.ruta_salida.mkdir(parents=True, exist_ok=True)
        self.ruta_csv = self._resolver_ruta_csv()            # ← nuevo

    def _resolver_ruta_csv(self) -> Path:
        """
        Detecta automáticamente el CSV de tickets más reciente.
        Acepta cualquier archivo con patrón: wispro_tickets_*.csv
        El más reciente por nombre (fecha en sufijo) tiene prioridad.
        """
        candidatos = sorted(
            self.ruta_entrada.glob("wispro_tickets_*.csv"),
            reverse=True
        )
        if candidatos:
            logger.info(f"CSV de tickets detectado: {candidatos[0].name}")
            return candidatos[0]

        # Fallback nombre fijo
        return self.ruta_entrada / "wispro_tickets.csv"

    # ------------------------------------------------------------------
    # BLOQUE 2: CARGA DEL ÍNDICE EMAIL → ID CUENTA
    # ------------------------------------------------------------------
    def _cargar_indice_email(self) -> dict:
        """
        Lee indice_email desde registro_procesados.json.
        Retorna dict {email_lower: id_cuenta}.
        """
        if not self.ruta_registro.exists():
            logger.warning(
                "registro_procesados.json no encontrado. "
                "Los tickets se generarán sin ID CUENTA."
            )
            return {}

        with open(self.ruta_registro, "r", encoding="utf-8") as f:
            data = json.load(f)

        indice = data.get("indice_email", {})
        logger.info(f"Índice email cargado: {len(indice)} entradas")
        return indice

    # ------------------------------------------------------------------
    # BLOQUE 3: LECTURA DEL CSV
    # ------------------------------------------------------------------
    def _leer_csv(self) -> pd.DataFrame:
        """
        Lee el CSV de tickets exportado desde Wispro.
        Normaliza columnas eliminando espacios extra.
        """
        if not self.ruta_csv.exists():
            raise FileNotFoundError(
                f"CSV de tickets no encontrado: {self.ruta_csv}\n"
                "Exporta desde Wispro → Mesa de ayuda → Exportar."
            )

        df = pd.read_csv(
            self.ruta_csv,
            encoding="utf-8",
            sep=",",
            dtype=str,
        ).fillna("")

        df.columns = [c.strip() for c in df.columns]
        logger.info(f"CSV cargado: {len(df)} tickets")
        return df

    # ------------------------------------------------------------------
    # BLOQUE 4: ENRIQUECIMIENTO (ID CUENTA + MES + CATEGORÍA)
    # ------------------------------------------------------------------
    def _enriquecer(self, df: pd.DataFrame, indice_email: dict) -> pd.DataFrame:
        """
        Agrega tres columnas calculadas:
        - id_cuenta   : cruzado por email desde el índice
        - mes_numero  : 1-24 según fecha de creación vs. inicio del contrato
        - categoria   : una de las 5 subcategorías contractuales
        Tickets sin ID CUENTA o fuera del rango 1-24 se incluyen con advertencia.
        """
        # ID CUENTA
        df["id_cuenta"] = df["Email"].str.strip().str.lower().map(
            lambda e: indice_email.get(e, "")
        )

        # Mes del contrato
        def calcular_mes(fecha_str: str) -> int:
            if not fecha_str:
                return 0
            match = re.search(r"(\d{2})/(\d{2})/(\d{4})", fecha_str)
            if not match:
                return 0
            dia, mes, anio = int(match.group(1)), int(match.group(2)), int(match.group(3))
            fecha = datetime(anio, mes, dia)
            delta = (fecha.year - CONTRATO_INICIO.year) * 12 + (fecha.month - CONTRATO_INICIO.month) + 1
            return delta if 1 <= delta <= CONTRATO_MESES else 0

        df["mes_numero"] = df["Creado el"].apply(calcular_mes)

        # Categoría contractual
        def resolver_categoria(cat: str) -> str:
            cat_norm = cat.strip().lower()
            return MAPA_CATEGORIA.get(cat_norm, "disponibilidad")  # default: disponibilidad

        df["categoria"] = df["Categoria"].apply(resolver_categoria)

        # Logs de advertencia
        sin_id  = df[df["id_cuenta"] == ""]
        fuera   = df[df["mes_numero"] == 0]
        if not sin_id.empty:
            logger.warning(f"{len(sin_id)} ticket(s) sin ID CUENTA")
        if not fuera.empty:
            logger.warning(f"{len(fuera)} ticket(s) fuera del rango de meses del contrato")

        return df

    # ------------------------------------------------------------------
    # BLOQUE 5: CONSTRUCCIÓN DE LA MATRIZ
    # ------------------------------------------------------------------
    def _construir_matriz(self, df: pd.DataFrame) -> dict:
        """
        Construye un dict anidado:
          matriz[id_cuenta][mes_numero][subcol] = [lista de N° ticket]
        Solo incluye tickets con id_cuenta y mes_numero válidos.
        """
        matriz = {}

        df_valido = df[(df["id_cuenta"] != "") & (df["mes_numero"] > 0)]

        for _, row in df_valido.iterrows():
            cuenta  = row["id_cuenta"]
            mes     = int(row["mes_numero"])
            cat     = row["categoria"]
            ticket  = row["Número del ticket"].strip()

            if cuenta not in matriz:
                matriz[cuenta] = {}
            if mes not in matriz[cuenta]:
                matriz[cuenta][mes] = {s: [] for s in SUBCOLS}

            matriz[cuenta][mes]["pqr"].append(ticket)
            matriz[cuenta][mes][cat].append(ticket)

        logger.info(f"Matriz construida: {len(matriz)} cuentas con tickets")
        return matriz

    # ------------------------------------------------------------------
    # BLOQUE 6: LISTA DE TODAS LAS CUENTAS REGISTRADAS
    # ------------------------------------------------------------------
    def _cargar_todas_las_cuentas(self) -> list:
        """
        Carga la lista completa de ID CUENTA desde registro_procesados.json
        para que TODAS las filas del contrato aparezcan en el Excel,
        incluso las que no tienen tickets ese periodo.
        """
        if not self.ruta_registro.exists():
            return []

        with open(self.ruta_registro, "r", encoding="utf-8") as f:
            data = json.load(f)

        # indice_email: {email: id_cuenta} → extraer valores únicos ordenados
        indice = data.get("indice_email", {})
        cuentas = sorted(set(indice.values()))
        logger.info(f"Total cuentas registradas: {len(cuentas)}")
        return cuentas

    # ------------------------------------------------------------------
    # BLOQUE 7: GENERACIÓN DEL EXCEL CONTRACTUAL
    # ------------------------------------------------------------------
    def _generar_excel(self, matriz: dict, todas_las_cuentas: list) -> Path:
        """
        Crea el Excel con el formato contractual:
        Filas 1-6  → Encabezado institucional
        Fila 7     → "CALIDAD DEL SERVICIO"
        Fila 8     → MES 1 | ... | MES 24  (celdas combinadas cada 5 cols)
        Fila 9     → ID CUENTA | PQR | Disp | Vel | Falla | Traslado | ...
        Fila 10+   → Datos por cuenta
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "PQRS"

        # ---- Estilos ----
        AZUL_HEADER  = "1F4E79"
        AZUL_MEDIO   = "2E75B6"
        AZUL_CLARO   = "DCE6F1"
        BLANCO       = "FFFFFF"
        GRIS_CLARO   = "F2F2F2"

        font_blanco_bold  = Font(bold=True, color="FFFFFF", size=9)
        font_negro_bold   = Font(bold=True, color="000000", size=9)
        font_normal       = Font(size=9)

        borde_fino = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        alin_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
        alin_izq    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

        # Total columnas: 1 (ID CUENTA) + 24 meses × 5 subcols
        total_cols = 1 + CONTRATO_MESES * len(SUBCOLS)

        # ---- FILAS 1-6: Encabezado institucional ----
        encabezados_inst = [
            f"RAZÓN SOCIAL DEL ISP: {ENCABEZADO_INFO['razon_social']}",
            f"MUNICIPIO: {ENCABEZADO_INFO['municipio']}",
            f"DEPARTAMENTO: {ENCABEZADO_INFO['departamento']}",
            f"CONTRATO DE FOMENTO No.: {ENCABEZADO_INFO['contrato']}",
            f"FECHA DEL CONTRATO DE FOMENTO: {ENCABEZADO_INFO['fecha']}",
        ]
        for i, texto in enumerate(encabezados_inst, start=1):
            ws.merge_cells(
                start_row=i, start_column=1,
                end_row=i,   end_column=total_cols,
            )
            celda = ws.cell(row=i, column=1, value=texto)
            celda.font      = font_negro_bold
            celda.alignment = alin_izq
            celda.border    = borde_fino
            ws.row_dimensions[i].height = 16

        # ---- FILA 6: "CALIDAD DEL SERVICIO" ----
        ws.merge_cells(
            start_row=6, start_column=1,
            end_row=6,   end_column=total_cols,
        )
        celda = ws.cell(row=6, column=1, value="CALIDAD DEL SERVICIO")
        celda.font      = Font(bold=True, color="FFFFFF", size=11)
        celda.fill      = PatternFill("solid", fgColor=AZUL_HEADER)
        celda.alignment = alin_centro
        celda.border    = borde_fino
        ws.row_dimensions[6].height = 20

        # ---- FILA 7: Cabecera de meses (MES 1 … MES 24) ----
        # Celda vacía bajo "ID CUENTA" (col 1)
        ws.merge_cells(start_row=7, start_column=1, end_row=8, end_column=1)
        celda_id = ws.cell(row=7, column=1, value="ID CUENTA\n(Número único)")
        celda_id.font      = font_blanco_bold
        celda_id.fill      = PatternFill("solid", fgColor=AZUL_HEADER)
        celda_id.alignment = alin_centro
        celda_id.border    = borde_fino

        for mes in range(1, CONTRATO_MESES + 1):
            col_inicio = 2 + (mes - 1) * len(SUBCOLS)
            col_fin    = col_inicio + len(SUBCOLS) - 1

            # Calcular nombre del mes (MES 1 = Dic 2025, MES 2 = Ene 2026 …)
            fecha_mes = datetime(
                CONTRATO_INICIO.year + (CONTRATO_INICIO.month + mes - 2) // 12,
                (CONTRATO_INICIO.month + mes - 2) % 12 + 1,
                1,
            )
            nombre_mes = fecha_mes.strftime("%b %Y").upper()
            etiqueta   = f"MES {mes}\n{nombre_mes}"

            ws.merge_cells(
                start_row=7, start_column=col_inicio,
                end_row=7,   end_column=col_fin,
            )
            celda = ws.cell(row=7, column=col_inicio, value=etiqueta)
            color = AZUL_HEADER if mes % 2 != 0 else AZUL_MEDIO
            celda.font      = font_blanco_bold
            celda.fill      = PatternFill("solid", fgColor=color)
            celda.alignment = alin_centro
            celda.border    = borde_fino

            # ---- FILA 8: Subcabeceras ----
            for j, subcol in enumerate(SUBCOLS):
                col = col_inicio + j
                c   = ws.cell(row=8, column=col, value=SUBCOLS_LABELS[subcol])
                c.font      = font_blanco_bold
                c.fill      = PatternFill("solid", fgColor=color)
                c.alignment = alin_centro
                c.border    = borde_fino
                ws.column_dimensions[get_column_letter(col)].width = (
                    18 if subcol == "pqr" else 14
                )

        ws.row_dimensions[7].height = 30
        ws.row_dimensions[8].height = 50

        # Ancho columna ID CUENTA
        ws.column_dimensions["A"].width = 16

        # ---- FILAS DE DATOS ----
        for fila_idx, cuenta in enumerate(todas_las_cuentas, start=9):
            color_fila = AZUL_CLARO if fila_idx % 2 == 0 else BLANCO

            # Columna ID CUENTA
            c = ws.cell(row=fila_idx, column=1, value=cuenta)
            c.font      = font_negro_bold
            c.fill      = PatternFill("solid", fgColor=color_fila)
            c.alignment = alin_centro
            c.border    = borde_fino

            # Rellenar cada mes
            for mes in range(1, CONTRATO_MESES + 1):
                col_inicio = 2 + (mes - 1) * len(SUBCOLS)
                datos_mes  = matriz.get(cuenta, {}).get(mes, {})

                for j, subcol in enumerate(SUBCOLS):
                    col    = col_inicio + j
                    valor  = ", ".join(datos_mes.get(subcol, [])) if datos_mes else ""
                    c      = ws.cell(row=fila_idx, column=col, value=valor)
                    c.font      = font_normal
                    c.fill      = PatternFill("solid", fgColor=color_fila)
                    c.alignment = alin_centro
                    c.border    = borde_fino

            ws.row_dimensions[fila_idx].height = 15

        # Congelar encabezados y primera columna
        ws.freeze_panes = ws.cell(row=9, column=2)

        # Guardar
        fecha_str  = datetime.now().strftime("%Y-%m-%d")
        nombre     = f"pqrs_calidad_servicio_{fecha_str}.xlsx"
        ruta_excel = self.ruta_salida / nombre
        wb.save(ruta_excel)

        logger.info(f"Excel generado: {ruta_excel}")
        return ruta_excel

    # ------------------------------------------------------------------
    # BLOQUE 8: MÉTODO PRINCIPAL
    # ------------------------------------------------------------------
    def generar(self) -> Path | None:
        """
        Orquesta la generación del Excel contractual de PQRS.
        Retorna None (sin error) si:
        - El CSV no existe → no hubo tickets esta semana
        - El CSV está vacío → ídem
        El informe semanal de instalaciones se genera independientemente.
        """
        # --- Guardia: CSV inexistente o vacío → semana sin PQRS ---
        if not self.ruta_csv.exists():
            logger.info(
                "CSV de tickets no encontrado — "
                "semana sin PQRS. No se genera Excel."
            )
            return None

        df_raw = self._leer_csv()

        if df_raw.empty:
            logger.info("CSV de tickets vacío — semana sin PQRS. No se genera Excel.")
            return None

        # --- Flujo normal ---
        logger.info(f"Iniciando generación PQRS contractual — {len(df_raw)} tickets...")

        indice_email   = self._cargar_indice_email()
        df_enriquecido = self._enriquecer(df_raw, indice_email)
        matriz         = self._construir_matriz(df_enriquecido)
        todas_cuentas  = self._cargar_todas_las_cuentas()

        cuentas_extra = sorted(set(matriz.keys()) - set(todas_cuentas))
        if cuentas_extra:
            logger.warning(
                f"{len(cuentas_extra)} cuenta(s) con tickets no encontradas "
                f"en el registro: {cuentas_extra}"
            )
        todas_cuentas = todas_cuentas + cuentas_extra

        ruta_excel = self._generar_excel(matriz, todas_cuentas)

        con_tickets = len([c for c in todas_cuentas if c in matriz])
        sin_tickets = len(todas_cuentas) - con_tickets

        logger.info(
            f"\n{'='*50}\n"
            f"  PQRS CALIDAD DEL SERVICIO GENERADO\n"
            f"  Total cuentas:     {len(todas_cuentas)}\n"
            f"  Con tickets:       {con_tickets}\n"
            f"  Sin tickets:       {sin_tickets}\n"
            f"  Tickets procesados:{len(df_raw)}\n"
            f"  Archivo:           {ruta_excel}\n"
            f"{'='*50}"
        )

        return ruta_excel


# ------------------------------------------------------------------
# Ejecución directa (pruebas)
# ------------------------------------------------------------------
def generar_pqrs():
    merger = TicketsMerger()
    merger.generar()


if __name__ == "__main__":
    generar_pqrs()
