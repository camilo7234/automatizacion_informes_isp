"""
Generador de reporte de facturación por cliente.

Objetivo:
- Construir un Excel independiente del informe semanal.
- Consolidar la información por cliente, incluso si tiene varios contratos.
- Relacionar facturas, contratos y clientes sin alterar el pipeline existente.
- Mostrar:
    * ID CUENTA (A 0000#) desde el índice de email del registro procesado
    * Datos del cliente
    * Contratos asociados
    * Total de facturas
    * Meses pagados
    * Último mes pagado
    * Primer vencimiento y segundo vencimiento de la última factura
    * Próximo corte calculado con la fecha más cercana al día operativo de corte
    * Estado financiero general

Notas:
- El reporte se guarda en: salidas/informes_facturacion/
- Este módulo NO modifica el informe semanal.
- Si una columna cambia de nombre en Wispro, el resolver de columnas intenta ubicarla por alias.
"""

from __future__ import annotations

import json
import logging
import re
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


# ------------------------------------------------------------------
# CONFIGURACIÓN GENERAL
# ------------------------------------------------------------------
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

BASE_ENTRADA = Path("datos/entrada/wispro")
BASE_SALIDA = Path("salidas/informes_facturacion")
RUTA_REGISTRO = Path("datos/procesados/modelo_contrato/registro_procesados.json")


# ------------------------------------------------------------------
# NORMALIZACIÓN Y UTILIDADES DE BÚSQUEDA (VERSIÓN ROBUSTA)
# ------------------------------------------------------------------
def _normalizar_texto(valor) -> str:
    """
    Convierte cualquier valor a un texto comparable:
    - elimina tildes
    - convierte a mayúsculas
    - elimina caracteres especiales
    - compacta espacios
    """
    if valor is None:
        return ""

    texto = str(valor).strip()
    if not texto:
        return ""

    texto = unicodedata.normalize("NFKD", texto)
    texto = texto.encode("ascii", "ignore").decode("ascii")
    texto = texto.upper()
    texto = re.sub(r"[^A-Z0-9]+", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def _normalizar_email(valor) -> str:
    if valor is None:
        return ""
    return str(valor).strip().lower()


def _normalizar_cedula(valor) -> str:
    """
    Normaliza una cédula eliminando el sufijo .0 que pandas genera
    al leer columnas numéricas como float.
        '1085896121.0'  →  '1085896121'
        '1085896121'    →  '1085896121'
        'nan'           →  ''
    """
    if valor is None:
        return ""
    s = str(valor).strip()
    if s.lower() in ("nan", "none", ""):
        return ""
    if s.endswith(".0"):
        s = s[:-2]
    return s


def _buscar_ultimo_archivo(carpeta: Path, patrones: list[str]) -> Path:
    """
    Busca el archivo más reciente.
    Incluye logging para trazabilidad.
    """
    if not carpeta.exists():
        raise FileNotFoundError(
            f"No existe la carpeta de entrada: {carpeta.resolve()}"
        )

    candidatos: list[Path] = []

    for patron in patrones:
        encontrados = list(carpeta.glob(patron))
        if encontrados:
            logger.info(f"Patrón '{patron}' → {len(encontrados)} archivo(s)")
        candidatos.extend(encontrados)

    if not candidatos:
        raise FileNotFoundError(
            f"No se encontró ningún archivo en {carpeta.resolve()} "
            f"con patrones: {patrones}"
        )

    candidatos.sort(key=lambda p: p.stat().st_mtime)
    seleccionado = candidatos[-1]

    logger.info(f"Archivo seleccionado: {seleccionado.name}")

    return seleccionado


def _leer_csv_robusto(ruta: Path) -> pd.DataFrame:
    """
    Lee CSV probando múltiples codificaciones.
    """
    errores = []

    for encoding in ("utf-8-sig", "utf-8", "latin1"):
        try:
            df = pd.read_csv(
                ruta,
                dtype=str,
                sep=None,
                engine="python",
                encoding=encoding,
            )
            logger.info(f"CSV leído correctamente con encoding: {encoding}")
            return df

        except Exception as exc:
            errores.append(f"{encoding}: {exc}")

    raise ValueError(
        f"No fue posible leer el CSV {ruta.resolve()}.\n"
        f"Intentos realizados:\n- " + "\n- ".join(errores)
    )


def _leer_excel_robusto(ruta: Path) -> pd.DataFrame:
    try:
        df = pd.read_excel(ruta, dtype=str)
        logger.info(f"Excel leído correctamente: {ruta.name}")
        return df
    except Exception as exc:
        raise ValueError(
            f"No fue posible leer el Excel {ruta.resolve()}.\n"
            f"Detalle técnico: {exc}"
        ) from exc


def _resolver_columna(df: pd.DataFrame, aliases: list[str], requerida: bool = False) -> str | None:
    """
    Busca una columna real en el DataFrame usando aliases.
    Ahora con logging y control.
    """
    mapa = {_normalizar_texto(col): col for col in df.columns}
    aliases_norm = [_normalizar_texto(alias) for alias in aliases]

    for alias_norm in aliases_norm:
        if alias_norm in mapa:
            col_encontrada = mapa[alias_norm]
            logger.info(f"Columna encontrada: {col_encontrada} (alias: {alias_norm})")
            return col_encontrada

    if requerida:
        raise KeyError(
            f"No se encontró columna requerida: {aliases}\n"
            f"Columnas disponibles: {list(df.columns)}"
        )

    logger.warning(f"No se encontró columna para aliases: {aliases}")
    return None


def _asegurar_columna_canonica(
    df: pd.DataFrame,
    nombre_canonico: str,
    aliases: list[str],
    requerida: bool = False,
) -> pd.DataFrame:
    """
    Garantiza columna estándar en el DataFrame.
    """
    col_real = _resolver_columna(df, aliases, requerida=requerida)

    if col_real is None:
        logger.warning(f"Columna '{nombre_canonico}' creada vacía")
        df[nombre_canonico] = ""
        return df

    if col_real != nombre_canonico:
        df[nombre_canonico] = df[col_real]

    df[nombre_canonico] = df[nombre_canonico].fillna("").astype(str).str.strip()

    return df


def _limpiar_dataframe_texto(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in df.columns:
        df[col] = df[col].fillna("").astype(str).str.strip()
    return df


# ------------------------------------------------------------------
# FECHAS, MONTO Y ESTADO DE FACTURACIÓN
# ------------------------------------------------------------------
def _parse_fecha(valor):
    """
    Convierte texto a Timestamp; si falla devuelve NaT.
    Se usa dayfirst=True porque Wispro y los CSV operativos suelen venir en formato dd/mm/yyyy.
    """
    if valor is None:
        return pd.NaT
    try:
        return pd.to_datetime(valor, dayfirst=True, errors="coerce")
    except Exception:
        return pd.NaT


def _parse_monto(valor):
    """
    Convierte montos tipo '$ 12.345,00', '12.345', '12345' o '12345.00' a float.
    Si no se puede interpretar, devuelve 0.0.
    """
    if valor is None:
        return 0.0

    texto = str(valor).strip()
    if not texto or texto.lower() in {"nan", "none"}:
        return 0.0

    # Limpieza de símbolos monetarios y separadores frecuentes.
    texto = texto.replace("$", "").replace("COP", "").replace(" ", "").strip()

    # Normalización para formatos latinos:
    # 12.345,67 -> 12345.67
    # 12,345.67 -> 12345.67 (se deja lo más compatible posible)
    if texto.count(",") == 1 and texto.count(".") >= 1:
        # Si hay coma decimal y puntos de miles
        texto = texto.replace(".", "").replace(",", ".")
    elif texto.count(",") == 1 and texto.count(".") == 0:
        # Decimal con coma
        texto = texto.replace(",", ".")
    else:
        # Caso simple o ya viene numérico con punto decimal
        texto = texto.replace(",", "")

    match = re.search(r"-?\d+(?:\.\d+)?", texto)
    if not match:
        return 0.0

    try:
        return float(match.group(0))
    except Exception:
        return 0.0


def _extraer_periodo(detalle) -> str | None:
    """
    Extrae el período mensual desde el campo DETALLES.
    Ejemplo:
        '2026-04-01 - 2026-04-30' -> '2026-04'
    """
    if not isinstance(detalle, str):
        return None

    detalle = detalle.strip()
    if not detalle:
        return None

    match = re.search(r"(\d{4}-\d{2})-\d{2}", detalle)
    return match.group(1) if match else None


def _clasificar_estado_factura(valor) -> bool | None:
    """
    Clasifica el estado de una factura:
    - True  -> pagada / cobrada
    - False -> pendiente / vencida / impaga
    - None  -> no se puede clasificar con certeza

    Se evita asumir de más: primero se excluyen estados claramente negativos.
    """
    texto = _normalizar_texto(valor)
    if not texto:
        return None

    negativos = ("IMPAG", "VENC", "PEND", "ANUL", "CANCEL", "MORA", "DEUDA")
    positivos = ("PAG", "COBR", "PAID", "SALD")

    if any(token in texto for token in negativos):
        return False

    if any(token in texto for token in positivos):
        return True

    return None


def _distancia_a_corte(fecha: pd.Timestamp) -> int:
    """
    Distancia en días al día 5 del mes de la fecha recibida.
    Se usa para seleccionar el vencimiento más cercano al corte operativo.
    """
    if pd.isna(fecha):
        return 999

    fecha = pd.Timestamp(fecha)
    fecha_corte = fecha.replace(day=5)
    return abs((fecha - fecha_corte).days)


def _seleccionar_fecha_principal(primer_v, segundo_v):
    """
    Selecciona la fecha de vencimiento más cercana al corte operativo del día 5.
    Si una de las fechas es inválida, toma la otra.
    Si ambas son inválidas, devuelve NaT.
    """
    fechas_validas = [f for f in [primer_v, segundo_v] if pd.notna(f)]
    if not fechas_validas:
        return pd.NaT
    if len(fechas_validas) == 1:
        return fechas_validas[0]
    return min(fechas_validas, key=_distancia_a_corte)


def _formatear_lista_unica(valores) -> str:
    """
    Convierte una serie/lista en texto único, conservando el orden de aparición
    y descartando vacíos.
    """
    vistos = set()
    salida = []

    for valor in valores:
        texto = str(valor).strip()
        if not texto or texto.lower() in {"nan", "none"}:
            continue
        if texto not in vistos:
            vistos.add(texto)
            salida.append(texto)

    return ", ".join(salida)

# ------------------------------------------------------------------
# FORMATO FINAL DE EXCEL (VERSIÓN ROBUSTA Y PROFESIONAL)
# ------------------------------------------------------------------
def _aplicar_formato_excel(ruta_xlsx: Path) -> None:
    """
    Aplica formato profesional al Excel:
    - encabezado estilizado
    - autofiltro
    - congelar primera fila
    - ajuste automático de columnas
    - formato de fechas
    - formato de moneda
    """
    try:
        from openpyxl.utils.datetime import from_excel

        wb = load_workbook(ruta_xlsx)

        if not wb.sheetnames:
            logger.warning("El archivo Excel no contiene hojas.")
            return

        ws = wb[wb.sheetnames[0]]

        if ws.max_row < 1 or ws.max_column < 1:
            logger.warning("La hoja está vacía, no se aplica formato.")
            return

        # ----------------------------------------------------------
        # CONGELAR PRIMERA FILA + FILTRO
        # ----------------------------------------------------------
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # ----------------------------------------------------------
        # ESTILO ENCABEZADO
        # ----------------------------------------------------------
        fill = PatternFill("solid", fgColor="1F4E78")
        font = Font(color="FFFFFF", bold=True)
        align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = align

        # ----------------------------------------------------------
        # AJUSTE DE COLUMNAS
        # ----------------------------------------------------------
        for col_idx in range(1, ws.max_column + 1):
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            max_length = 0

            for row_idx in range(1, ws.max_row + 1):
                try:
                    value = ws.cell(row=row_idx, column=col_idx).value
                    if value not in (None, ""):
                        max_length = max(max_length, len(str(value)))
                except Exception:
                    pass

            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # ----------------------------------------------------------
        # FORMATO DE FECHAS Y MONEDA
        # ----------------------------------------------------------
        for col_idx in range(1, ws.max_column + 1):
            header = str(ws.cell(row=1, column=col_idx).value or "").upper()

            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)

                if cell.value in (None, ""):
                    continue

                # ---- FECHAS ----
                if "FECHA" in header or "VENCIMIENTO" in header or "CORTE" in header:
                    try:
                        valor = cell.value

                        # Si ya viene como fecha real
                        if hasattr(valor, "strftime"):
                            cell.number_format = "DD/MM/YYYY"
                            cell.alignment = Alignment(horizontal="center")
                            continue

                        # Si viene como serial numérico de Excel
                        if isinstance(valor, (int, float)):
                            # Seriales pequeños como 96 o 51 suelen ser basura en este contexto.
                            # Los dejamos vacíos para no mostrar fechas absurdas de 1900.
                            if float(valor) < 10000:
                                cell.value = ""
                                continue

                            fecha = from_excel(float(valor))
                            cell.value = fecha
                            cell.number_format = "DD/MM/YYYY"
                            cell.alignment = Alignment(horizontal="center")
                            continue

                        # Si viene como texto
                        if isinstance(valor, str):
                            texto = valor.strip()

                            # Si el texto es solo número, no lo mostramos como fecha.
                            if re.fullmatch(r"\d+(\.\d+)?", texto):
                                cell.value = ""
                                continue

                            parsed = pd.to_datetime(texto, errors="coerce", dayfirst=True)
                            if pd.notna(parsed):
                                cell.value = parsed.to_pydatetime()
                                cell.number_format = "DD/MM/YYYY"
                                cell.alignment = Alignment(horizontal="center")
                            else:
                                # Si no se puede interpretar como fecha, se limpia
                                cell.value = ""
                            continue

                    except Exception:
                        # Si algo falla en una celda de fecha, la dejamos vacía
                        cell.value = ""

                # ---- MONEDA ----
                if "VALOR" in header or "TOTAL PAGADO" in header:
                    try:
                        valor = cell.value
                        if isinstance(valor, str):
                            texto = valor.replace("$", "").replace("COP", "").replace(" ", "").strip()
                            if texto.count(",") == 1 and texto.count(".") >= 1:
                                texto = texto.replace(".", "").replace(",", ".")
                            elif texto.count(",") == 1 and texto.count(".") == 0:
                                texto = texto.replace(",", ".")
                            else:
                                texto = texto.replace(",", "")

                            match = re.search(r"-?\d+(?:\.\d+)?", texto)
                            if match:
                                cell.value = float(match.group(0))
                            else:
                                cell.value = 0.0

                        elif isinstance(valor, (int, float)):
                            cell.value = float(valor)
                        else:
                            cell.value = 0.0

                        cell.number_format = '"$"#,##0.00'
                        cell.alignment = Alignment(horizontal="right")
                    except Exception:
                        pass

        # ----------------------------------------------------------
        # GUARDAR
        # ----------------------------------------------------------
        wb.save(ruta_xlsx)

        logger.info("Formato Excel aplicado correctamente.")

    except Exception as exc:
        logger.warning(f"No fue posible aplicar formato avanzado al Excel: {exc}")

        
# ------------------------------------------------------------------
# GENERADOR PRINCIPAL
# ------------------------------------------------------------------
def generar_reporte_facturacion() -> Path:
    """
    Genera el reporte de facturación por cliente y lo guarda en:
        salidas/informes_facturacion/reporte_facturacion_YYYY-MM-DD.xlsx

    Retorna:
        Path del archivo generado.
    """
    logger.info("Generando reporte de facturación por clientes...")

    # --------------------------------------------------------------
    # RUTAS DE ENTRADA / SALIDA
    # --------------------------------------------------------------
    BASE_SALIDA.mkdir(parents=True, exist_ok=True)

    ruta_clientes = _buscar_ultimo_archivo(
        BASE_ENTRADA,
        ["wispro_clientes_*.csv", "clientes_*.csv"]
    )
    ruta_contratos = _buscar_ultimo_archivo(
        BASE_ENTRADA,
        ["wispro_contratos_*.xls", "wispro_contratos_*.xlsx", "wispro_contratos_*.csv" ,"contratos_*.xls", "contratos_*.csv","contratos_*.xlsx"]
    )
    ruta_facturas = _buscar_ultimo_archivo(
        BASE_ENTRADA,
        ["wispro_facturas_*.csv", "facturas_*.csv"]
    )

    if not RUTA_REGISTRO.exists():
        raise FileNotFoundError(
            f"No existe el archivo de registro: {RUTA_REGISTRO.resolve()}"
        )

    logger.info(f"Clientes:  {ruta_clientes.name}")
    logger.info(f"Contratos: {ruta_contratos.name}")
    logger.info(f"Facturas:  {ruta_facturas.name}")

    # --------------------------------------------------------------
    # CARGA DE ARCHIVOS (ROBUSTA — SOPORTA CSV Y EXCEL)
    # --------------------------------------------------------------

    # -------------------------
    # CLIENTES (SIEMPRE CSV EN TU CASO)
    # -------------------------
    df_clientes = _leer_csv_robusto(ruta_clientes)

    # -------------------------
    # CONTRATOS (PUEDE SER CSV O EXCEL)
    # -------------------------
    if ruta_contratos.suffix.lower() == ".csv":
        logger.info(f"Leyendo contratos como CSV: {ruta_contratos.name}")
        df_contratos = _leer_csv_robusto(ruta_contratos)
    else:
        logger.info(f"Leyendo contratos como Excel: {ruta_contratos.name}")
        df_contratos = _leer_excel_robusto(ruta_contratos)

    # -------------------------
    # FACTURAS (CSV)
    # -------------------------
    df_facturas = _leer_csv_robusto(ruta_facturas)

    # -------------------------
    # REGISTRO JSON (ID CUENTA)
    # -------------------------
    if not RUTA_REGISTRO.exists():
        raise FileNotFoundError(
            f"No se encontró el archivo de registro: {RUTA_REGISTRO.resolve()}"
        )

    with open(RUTA_REGISTRO, "r", encoding="utf-8") as f:
        registro = json.load(f)

    # Normalización del índice de emails → ID CUENTA
    indice_email = {
        _normalizar_email(k): v
        for k, v in registro.get("indice_email", {}).items()
    }

    logger.info(f"Índice de emails cargado: {len(indice_email)} registros")

    # --------------------------------------------------------------
    # LIMPIEZA DE DATOS Y NORMALIZACIÓN DE COLUMNAS
    # --------------------------------------------------------------
    df_clientes = _limpiar_dataframe_texto(df_clientes)
    df_contratos = _limpiar_dataframe_texto(df_contratos)
    df_facturas = _limpiar_dataframe_texto(df_facturas)

    # --- CLIENTES ---
    df_clientes = _asegurar_columna_canonica(df_clientes, "ID CLIENTE", ["ID CLIENTE", "ID_CLIENTE"], requerida=True)
    df_clientes = _asegurar_columna_canonica(df_clientes, "ID PERSONALIZABLE", ["ID PERSONALIZABLE", "ID PERSONALIZABLE_cliente"], requerida=False)
    df_clientes = _asegurar_columna_canonica(df_clientes, "NOMBRE", ["NOMBRE", "NOMBRE CLIENTE", "NOMBRE COMPLETO"], requerida=True)
    df_clientes = _asegurar_columna_canonica(df_clientes, "EMAIL", ["EMAIL", "EMAIL_cliente"], requerida=False)
    df_clientes = _asegurar_columna_canonica(df_clientes, "TELÉFONO", ["TELÉFONO", "TELÉFONO CELULAR", "TELÉFONOS"], requerida=False)
    df_clientes = _asegurar_columna_canonica(df_clientes, "DIRECCIÓN", ["DIRECCIÓN", "DIRECCIÓN DEL CLIENTE", "DIRRECIÓN DEL CLIENTE"], requerida=False)
    df_clientes = _asegurar_columna_canonica(df_clientes, "DOCUMENTO/CÉDULA", ["DOCUMENTO/CÉDULA", "DOCUMENTO O CÉDULA DE IDENTIDAD CLIENTE", "DOCUMENTO O CEDULA DE IDENTIDAD CLIENTE"], requerida=False)

    # --- CONTRATOS ---
    df_contratos = _asegurar_columna_canonica(df_contratos, "ID CONTRATO", ["ID CONTRATO", "ID"], requerida=True)
    df_contratos = _asegurar_columna_canonica(df_contratos, "ID CLIENTE", ["ID CLIENTE", "ID CLIENTE_contrato", "ID CLIENTE_cliente"], requerida=False)
    df_contratos = _asegurar_columna_canonica(df_contratos, "ESTADO CONTRATO", ["ESTADO_contrato", "ESTADO", "ESTADO DEL CONTRATO"], requerida=False)
    df_contratos = _asegurar_columna_canonica(df_contratos, "PLAN", ["NOMBRE PLAN_contrato", "NOMBRE PLAN", "PLAN"], requerida=False)
    df_contratos = _asegurar_columna_canonica(df_contratos, "FECHA DE ALTA", ["FECHA DE ALTA", "CREADO EL_contrato", "CREADO EL"], requerida=False)

    # [C-02] Filtrar contratos deshabilitados.
    # Wispro marca contratos suspendidos como "Deshabilitado".
    # Esos contratos no deben entrar al reporte como activos ni generar facturación.
    if "ESTADO CONTRATO" in df_contratos.columns:
        antes_contratos = len(df_contratos)
        df_contratos = df_contratos[
            df_contratos["ESTADO CONTRATO"].str.strip().str.upper() == "HABILITADO"
        ].copy()
        logger.info(
            f"[C-02] Contratos filtrados por ESTADO CONTRATO=HABILITADO: "
            f"{antes_contratos} → {len(df_contratos)}"
        )
    else:
        logger.warning("[C-02] No se encontró columna 'ESTADO CONTRATO'; no se filtraron contratos deshabilitados.")

    # --- FACTURAS ---
    df_facturas = _asegurar_columna_canonica(df_facturas, "ID CLIENTE", ["ID CLIENTE", "ID_CLIENTE"], requerida=True)
    df_facturas = _asegurar_columna_canonica(df_facturas, "ID CONTRATO", ["ID CONTRATO", "ID_CONTRATO"], requerida=True)
    df_facturas = _asegurar_columna_canonica(df_facturas, "NOMBRE CLIENTE", ["NOMBRE CLIENTE", "NOMBRE CLIENTE_contrato"], requerida=False)
    df_facturas = _asegurar_columna_canonica(df_facturas, "PRIMER VENCIMIENTO", ["PRIMER VENCIMIENTO"], requerida=False)
    df_facturas = _asegurar_columna_canonica(df_facturas, "SEGUNDO VENCIMIENTO", ["SEGUNDO VENCIMIENTO"], requerida=False)
    df_facturas = _asegurar_columna_canonica(df_facturas, "DETALLES", ["DETALLES"], requerida=False)
    df_facturas = _asegurar_columna_canonica(df_facturas, "ESTADO FACTURA", ["ESTADO", "ESTADO FACTURA", "ESTADO_pago"], requerida=False)
    df_facturas = _asegurar_columna_canonica(df_facturas, "MONTO", ["MONTO", "VALOR", "TOTAL"], requerida=False)
    df_facturas = _asegurar_columna_canonica(df_facturas, "FECHA EMISIÓN", ["FECHA EMISIÓN", "FECHA DE EMISIÓN", "EMITIDA EL", "CREADO EL"], requerida=False)

    # --------------------------------------------------------------
    # TIPADO Y CAMPOS CALCULADOS
    # --------------------------------------------------------------
    df_facturas["PRIMER VENCIMIENTO"] = df_facturas["PRIMER VENCIMIENTO"].apply(_parse_fecha)
    df_facturas["SEGUNDO VENCIMIENTO"] = df_facturas["SEGUNDO VENCIMIENTO"].apply(_parse_fecha)
    df_facturas["FECHA EMISIÓN"] = df_facturas["FECHA EMISIÓN"].apply(_parse_fecha)

    # Período facturado (mes) a partir del detalle; si no existe, usa fecha de emisión.
    df_facturas["PERIODO_FACTURADO"] = df_facturas["DETALLES"].apply(_extraer_periodo)
    mask_sin_periodo = df_facturas["PERIODO_FACTURADO"].isna() | (df_facturas["PERIODO_FACTURADO"].astype(str).str.strip() == "")
    df_facturas.loc[mask_sin_periodo, "PERIODO_FACTURADO"] = (
        df_facturas.loc[mask_sin_periodo, "FECHA EMISIÓN"]
        .dt.strftime("%Y-%m")
    )

    # Valor numérico para poder sumar correctamente.
    df_facturas["MONTO_NUM"] = df_facturas["MONTO"].apply(_parse_monto)

    # [C-03] BALANCE_NUM: monto pendiente según Wispro.
    # Wispro ya trae una columna BALANCE que indica deuda de la factura.
    #  - 0.0     → factura saldada
    #  - > 0.0   → monto pendiente
    # La convertimos a float para poder sumar por cliente.
    if "BALANCE" in df_facturas.columns:
        df_facturas["BALANCE_NUM"] = df_facturas["BALANCE"].apply(_parse_monto)
    else:
        df_facturas["BALANCE_NUM"] = 0.0
        logger.warning("[C-03] No se encontró columna BALANCE en facturas; BALANCE_NUM se deja en 0.0")

    # [C-01] Excluir facturas ANULADAS antes de cualquier cálculo.
    # Wispro marca como "Anulado" registros de prueba y reversiones.
    # No representan ni deuda ni pago real — contaminarían totales.
    _antes_anulados = len(df_facturas)
    df_facturas = df_facturas[
        df_facturas["ESTADO FACTURA"].str.strip().str.upper() != "ANULADO"
    ].copy()
    _excluidas = _antes_anulados - len(df_facturas)
    if _excluidas > 0:
        logger.info(f"[C-01] Facturas anuladas excluidas: {_excluidas} (quedan {len(df_facturas)})")

    # [C-01] ES_PAGADA tomado directamente del campo ESTADO de Wispro.
    # Se eliminó _clasificar_estado_factura() porque recalculaba
    # lo que Wispro ya resolvió con criterio fiscal.
    # "Pagado" = pagada. Cualquier otro valor = no pagada.
    df_facturas["ES_PAGADA"] = (
        df_facturas["ESTADO FACTURA"].str.strip().str.upper() == "PAGADO"
    )
    logger.info(
        f"[C-01] Estado facturas → Pagadas: {df_facturas['ES_PAGADA'].sum()} | "
        f"Impagas: {(~df_facturas['ES_PAGADA']).sum()}"
    )

    # Fecha de referencia para detectar la factura más reciente por cliente.
    # Se prioriza la fecha de emisión; si falta, se usa el primer y luego el segundo vencimiento.
    df_facturas["_FECHA_REFERENCIA"] = df_facturas["FECHA EMISIÓN"]
    mask_ref = df_facturas["_FECHA_REFERENCIA"].isna()
    df_facturas.loc[mask_ref, "_FECHA_REFERENCIA"] = df_facturas.loc[mask_ref, "PRIMER VENCIMIENTO"]
    mask_ref = df_facturas["_FECHA_REFERENCIA"].isna()
    df_facturas.loc[mask_ref, "_FECHA_REFERENCIA"] = df_facturas.loc[mask_ref, "SEGUNDO VENCIMIENTO"]

    # --------------------------------------------------------------
    # DEPURACIÓN DE DUPLICADOS
    # --------------------------------------------------------------
    subset_dedup = [col for col in ["ID CLIENTE", "ID CONTRATO", "PERIODO_FACTURADO", "PRIMER VENCIMIENTO", "SEGUNDO VENCIMIENTO"] if col in df_facturas.columns]
    if subset_dedup:
        df_facturas = df_facturas.drop_duplicates(subset=subset_dedup, keep="last")

    df_contratos = df_contratos.drop_duplicates(subset=["ID CONTRATO"], keep="last")
    df_clientes = df_clientes.drop_duplicates(subset=["ID CLIENTE"], keep="last")

    # --------------------------------------------------------------
    # MERGE: FACTURAS + CONTRATOS + CLIENTES
    # --------------------------------------------------------------
    # Merge por contrato
    df = df_facturas.merge(
        df_contratos[
            [col for col in df_contratos.columns if col in {
                "ID CONTRATO",
                "ID CLIENTE",
                "ESTADO CONTRATO",
                "PLAN",
                "FECHA DE ALTA",
            }]
        ],
        on="ID CONTRATO",
        how="left",
        suffixes=("", "_CONTRATO"),
    )

    # Merge por cliente
    df = df.merge(
        df_clientes[
            [col for col in df_clientes.columns if col in {
                "ID CLIENTE",
                "ID PERSONALIZABLE",
                "NOMBRE",
                "EMAIL",
                "TELÉFONO",
                "DIRECCIÓN",
                "DOCUMENTO/CÉDULA",
                # [C-05] traemos también TIPO DE FACTURA si existe en clientes
                "TIPO DE FACTURA",
            } if col in df_clientes.columns]
        ],
        on="ID CLIENTE",
        how="left",
        suffixes=("", "_CLIENTE"),
    )

    # [C-05] Normalizar tipo de facturación (Factura de Venta vs Comprobante/Fantasia)
    #  - Si facturas.csv trae TIPO FACTURA, se usa como fuente principal.
    #  - Si no, se usa TIPO DE FACTURA de clientes.csv.
    #  - Resultado se deja en TIPO_FACTURACION para poder filtrar o segmentar después.
    tipo_factura_facturas = df_facturas.columns[df_facturas.columns.str.upper().str.contains("TIPO FACTURA")]
    col_tipo_facturas = tipo_factura_facturas[0] if len(tipo_factura_facturas) > 0 else None

    if col_tipo_facturas and col_tipo_facturas in df.columns:
        df["TIPO_FACTURACION"] = df[col_tipo_facturas].fillna("").astype(str).str.strip()
    elif "TIPO DE FACTURA" in df.columns:
        df["TIPO_FACTURACION"] = df["TIPO DE FACTURA"].fillna("").astype(str).str.strip()
    else:
        df["TIPO_FACTURACION"] = ""
        logger.warning("[C-05] No se encontró TIPO FACTURA ni TIPO DE FACTURA; TIPO_FACTURACION queda vacío.")

    # Normalización de claves para evitar blancos
    df["ID CLIENTE"] = df["ID CLIENTE"].fillna("").astype(str).str.strip()
    df["ID CONTRATO"] = df["ID CONTRATO"].fillna("").astype(str).str.strip()

    df = df[df["ID CLIENTE"] != ""].copy()

    # --------------------------------------------------------------
    # (C-SNAPSHOT) MÓDULO DE FACTURACIÓN → SIEMPRE SNAPSHOT COMPLETO
    # --------------------------------------------------------------
    # A diferencia del informe semanal, aquí NO se usa el mecanismo
    # de anti-duplicación por cédula (cedulas_procesadas), porque:
    #   - El estado de pago de cualquier cliente puede cambiar semana a semana.
    #   - Necesitamos siempre la foto completa de todos los clientes
    #     con la información MÁS RECIENTE.
    #   - Si filtráramos por cedulas_procesadas, un cliente que estaba
    #     en mora y luego paga nunca actualizaría su estado en este reporte.
    #
    # Por compatibilidad futura, solo dejamos un log informativo con
    # la cantidad de cédulas registradas, pero NO filtramos ni cortamos
    # el DataFrame df en este módulo.
    try:
        cedulas_ya_procesadas = set()
        with open(RUTA_REGISTRO, "r", encoding="utf-8") as f:
            _reg = json.load(f)
        for c in _reg.get("cedulas_procesadas", []):
            cedulas_ya_procesadas.add(_normalizar_cedula(c))
        logger.info(
            f"[SNAPSHOT] Registro de cedulas_procesadas cargado "
            f"({len(cedulas_ya_procesadas)} cédulas), "
            "pero NO se aplica filtro en reporte de facturación."
        )
    except Exception as e:
        logger.warning(
            f"[SNAPSHOT] No se pudo cargar cedulas_procesadas: {e}. "
            "Se continua sin filtro (comportamiento esperado en facturación)."
        )

    if df.empty:
        raise ValueError(
            "Después de cruzar facturas, contratos y clientes no quedaron registros válidos.\n"
            "Revisa que los archivos de Wispro tengan coincidentes los campos ID CLIENTE e ID CONTRATO."
        )
    # --------------------------------------------------------------
    # AGRUPACIÓN POR CLIENTE (VERSIÓN ROBUSTA)
    # --------------------------------------------------------------
    resultados = []

    def _primer_valor_valido(*valores):
        """
        Devuelve el primer valor útil entre varios candidatos.
        Ignora None, NaN, cadenas vacías y el texto literal 'nan'.
        """
        for valor in valores:
            if valor is None:
                continue
            if pd.isna(valor):
                continue

            texto = str(valor).strip()
            if not texto or texto.lower() in {"nan", "none"}:
                continue

            return texto
        return ""

    for cliente_id, grupo in df.groupby("ID CLIENTE", dropna=False):
        grupo = grupo.copy()

        # ----------------------------------------------------------
        # ORDENAMIENTO SEGURO (última factura real)
        # ----------------------------------------------------------
        grupo["_FECHA_ORDEN"] = grupo["_FECHA_REFERENCIA"]

        grupo["_FECHA_ORDEN"] = grupo["_FECHA_ORDEN"].fillna(grupo["SEGUNDO VENCIMIENTO"])
        grupo["_FECHA_ORDEN"] = grupo["_FECHA_ORDEN"].fillna(grupo["PRIMER VENCIMIENTO"])

        grupo = grupo.sort_values(
            by=["_FECHA_ORDEN"],
            na_position="last"
        )

        ultimo = grupo.iloc[-1]

        # ----------------------------------------------------------
        # IDENTIFICACIÓN DE LA CUENTA (A0000#)
        # Ruta 1: buscar por email en indice_email
        # ✅ CORRECCIÓN #2 — Ruta 2 (fallback): buscar por cédula
        # en indice_cedula si la Ruta 1 no encontró coincidencia.
        # ----------------------------------------------------------

        # --- Cargar índice de cédulas (una sola vez fuera del loop
        #     sería más eficiente, pero aquí se mantiene el patrón
        #     existente del módulo para no alterar más estructura) ---
        indice_cedula = {
            str(k).strip(): v
            for k, v in registro.get("indice_cedula", {}).items()
        }

        # RUTA 1 — búsqueda por email
        emails_candidatos = []
        for col in ("EMAIL", "EMAIL_CLIENTE", "EMAIL_contrato"):
            if col in grupo.columns:
                for valor in grupo[col].tolist():
                    email_limpio = _normalizar_email(valor)
                    if email_limpio and email_limpio not in {"nan", "none"}:
                        if email_limpio not in emails_candidatos:
                            emails_candidatos.append(email_limpio)

        email = ""
        for candidato in emails_candidatos:
            if candidato in indice_email:
                email = candidato
                break

        # Fallback email: última fila del grupo
        if not email:
            email_ultimo = _primer_valor_valido(
                ultimo.get("EMAIL", None),
                ultimo.get("EMAIL_CLIENTE", None),
                ultimo.get("EMAIL_contrato", None),
            )
            email_ultimo = _normalizar_email(email_ultimo)
            if email_ultimo in indice_email:
                email = email_ultimo

        id_cuenta = indice_email.get(email, "") if email else ""

        # ✅ CORRECCIÓN BUG #2 — RUTA 2: fallback por cédula
        if not id_cuenta:
            documento_raw = _primer_valor_valido(
                ultimo.get("DOCUMENTO/CÉDULA", None),
                ultimo.get("DOCUMENTO/CEDULA", None),
            )
            cedula_norm = _normalizar_cedula(documento_raw)

            # Buscar en todas las cédulas del grupo, no solo la última fila
            cedulas_grupo = []
            for col in ("DOCUMENTO/CÉDULA", "DOCUMENTO/CEDULA"):
                if col in grupo.columns:
                    for val in grupo[col].tolist():
                        c = _normalizar_cedula(val)
                        if c and c not in {"nan", "none"} and c not in cedulas_grupo:
                            cedulas_grupo.append(c)

            for cedula_candidata in cedulas_grupo:
                if cedula_candidata in indice_cedula:
                    id_cuenta = indice_cedula[cedula_candidata]
                    logger.info(
                        f"ID CUENTA resuelto por cédula ({cedula_candidata}): {id_cuenta}"
                    )
                    break

        # Si ninguna ruta encontró el ID, marcar como SIN_ID
        if not id_cuenta:
            id_cuenta = "SIN_ID"
            logger.warning(
                f"No se encontró ID CUENTA para cliente {cliente_id}. "
                f"Emails intentados: {emails_candidatos}. "
                f"Cédulas intentadas: {cedulas_grupo if 'cedulas_grupo' in dir() else '—'}"
            )

        documento = _primer_valor_valido(
            ultimo.get("DOCUMENTO/CÉDULA", None),
            ultimo.get("DOCUMENTO/CEDULA", None),
        )

        # ----------------------------------------------------------
        # CONTRATOS ASOCIADOS
        # ----------------------------------------------------------
        contratos_asociados = _formatear_lista_unica(grupo["ID CONTRATO"].tolist())

        estados_contrato = _formatear_lista_unica(
            grupo.get("ESTADO CONTRATO", pd.Series(dtype=str)).tolist()
        )

        planes_asociados = _formatear_lista_unica(
            grupo.get("PLAN", pd.Series(dtype=str)).tolist()
        )

        # ----------------------------------------------------------
        # DATOS DE FACTURA MÁS RECIENTE
        # ----------------------------------------------------------
        primer_v = ultimo.get("PRIMER VENCIMIENTO", pd.NaT)
        segundo_v = ultimo.get("SEGUNDO VENCIMIENTO", pd.NaT)
        fecha_emision_ultima = ultimo.get("FECHA EMISIÓN", pd.NaT)

        # Si la última fila no trae fechas válidas, se toma la última fecha válida del grupo.
        if pd.isna(primer_v) and "PRIMER VENCIMIENTO" in grupo.columns:
            validos_primer = grupo["PRIMER VENCIMIENTO"].dropna()
            if not validos_primer.empty:
                primer_v = validos_primer.iloc[-1]

        if pd.isna(segundo_v) and "SEGUNDO VENCIMIENTO" in grupo.columns:
            validos_segundo = grupo["SEGUNDO VENCIMIENTO"].dropna()
            if not validos_segundo.empty:
                segundo_v = validos_segundo.iloc[-1]

        if pd.isna(fecha_emision_ultima) and "FECHA EMISIÓN" in grupo.columns:
            validos_emision = grupo["FECHA EMISIÓN"].dropna()
            if not validos_emision.empty:
                fecha_emision_ultima = validos_emision.iloc[-1]

        proximo_corte = _seleccionar_fecha_principal(primer_v, segundo_v)

        # ----------------------------------------------------------
        # FACTURACIÓN (ROBUSTA)
        # ----------------------------------------------------------
        # [C-01] Solo facturas donde Wispro confirmó ESTADO="Pagado".
        # Se eliminó el OR con MONTO_NUM > 0 porque las facturas IMPAGAS
        # también tienen monto > 0, lo que inflaba FACTURAS PAGADAS
        # y VALOR TOTAL PAGADO con deudas reales del cliente.
        grupo_pagado = grupo[grupo["ES_PAGADA"] == True].copy()

        facturas_pagadas = len(grupo_pagado)
        facturas_total = len(grupo)

        meses_pagados = sorted(
            set(
                x for x in grupo_pagado["PERIODO_FACTURADO"].tolist()
                if isinstance(x, str) and x.strip()
            )
        )

        ultimo_mes_pagado = meses_pagados[-1] if meses_pagados else ""

        total_pagado = float(
            grupo_pagado["MONTO_NUM"].fillna(0).sum()
        ) if facturas_pagadas else 0.0

        # ----------------------------------------------------------
        # ESTADO GENERAL (AJUSTADO A OPERACIÓN REAL ISP)
        # ----------------------------------------------------------
        # [C-01] El estado se determina primero por ES_PAGADA (Wispro).
        # Antes un cliente que pagó a tiempo pero con fecha de vencimiento
        # ya superada aparecía como "EN MORA" — incorrecto.
        # Ahora: si Wispro dice PAGADO → AL DÍA, sin importar la fecha.
        # Solo se evalúan fechas cuando la factura más reciente es IMPAGA.
        hoy = pd.Timestamp.today().normalize()
        es_pagada_ultimo = bool(ultimo.get("ES_PAGADA", False))

        if es_pagada_ultimo:
            estado_general = "AL DÍA"

        elif pd.notna(segundo_v):
            segundo_v_norm = pd.Timestamp(segundo_v).normalize()
            fecha_corte_real = segundo_v_norm + pd.Timedelta(days=5)

            if hoy > fecha_corte_real:
                estado_general = "EN MORA"
            elif hoy > segundo_v_norm:
                estado_general = "ALERTA"
            else:
                estado_general = "PENDIENTE"

        elif pd.notna(primer_v):
            primer_v_norm = pd.Timestamp(primer_v).normalize()

            if hoy > primer_v_norm:
                estado_general = "ALERTA"
            else:
                estado_general = "PENDIENTE"

        else:
            estado_general = "SIN FECHA"


        # ----------------------------------------------------------
        # RESULTADO FINAL POR CLIENTE
        # ----------------------------------------------------------
        # [C-03] Se agrega DEUDA_TOTAL basada en BALANCE_NUM de Wispro.
        # BALANCE_NUM es la suma de saldos pendientes de todas las facturas
        # del cliente (después de excluir ANULADAS).
        deuda_total = float(
            grupo["BALANCE_NUM"].fillna(0).sum()
        ) if "BALANCE_NUM" in grupo.columns else 0.0

        # [C-06] Datos de morosidad que ya trae Wispro a nivel de cliente.
        # NÚMERO DE FACTURAS IMPAGAS y BALANCE DE FACTRAS IMPAGAS
        # vienen de wispro_clientes.csv. Si faltan, se asumen 0.
        num_impagas_cliente = 0
        balance_impagas_cliente = 0.0

        if "NÚMERO DE FACTURAS IMPAGAS" in grupo.columns:
            try:
                # Tomamos el valor de la última fila no vacía del grupo
                serie_num = grupo["NÚMERO DE FACTURAS IMPAGAS"].dropna()
                if not serie_num.empty:
                    num_impagas_cliente = int(float(serie_num.iloc[-1]))  # maneja "12.0"
            except Exception:
                num_impagas_cliente = 0

        if "BALANCE DE FACTRAS IMPAGAS" in grupo.columns:
            try:
                serie_bal = grupo["BALANCE DE FACTRAS IMPAGAS"].dropna()
                if not serie_bal.empty:
                    balance_impagas_cliente = float(_parse_monto(serie_bal.iloc[-1]))
            except Exception:
                balance_impagas_cliente = 0.0

        resultados.append({
            "ID CUENTA": id_cuenta,
            "ID CLIENTE": cliente_id,
            "NOMBRE": str(ultimo.get("NOMBRE", "")).strip(),
            "DOCUMENTO": documento,
            "EMAIL": email,
            "TELÉFONO": str(ultimo.get("TELÉFONO", "")).strip(),
            "DIRECCIÓN": str(ultimo.get("DIRECCIÓN", "")).strip(),

            "ID CONTRATO PRINCIPAL": str(ultimo.get("ID CONTRATO", "")).strip(),
            "CONTRATOS ASOCIADOS": contratos_asociados,
            "ESTADOS CONTRATO": estados_contrato,
            "PLANES ASOCIADOS": planes_asociados,

            "TOTAL FACTURAS": facturas_total,
            "FACTURAS PAGADAS": facturas_pagadas,
            "MESES PAGADOS": len(meses_pagados),
            "ÚLTIMO MES PAGADO": ultimo_mes_pagado,
            "VALOR TOTAL PAGADO": total_pagado,
            "DEUDA_TOTAL": deuda_total,                  # [C-03] saldo pendiente según Wispro (facturas)
            "FACTURAS_IMPAGAS_CLIENTE": num_impagas_cliente,      # [C-06]
            "BALANCE_IMPAGAS_CLIENTE": balance_impagas_cliente,   # [C-06]

            "FECHA EMISIÓN (FACTURA ÚLTIMA)": fecha_emision_ultima,
            "PRIMER VENCIMIENTO": primer_v,
            "SEGUNDO VENCIMIENTO": segundo_v,
            "PRÓXIMO CORTE": proximo_corte,

            "ESTADO GENERAL": estado_general,
        })


    # --------------------------------------------------------------
    # DATAFRAME FINAL Y EXPORTACIÓN (CORREGIDO Y ROBUSTO)
    # --------------------------------------------------------------
    df_final = pd.DataFrame(resultados)

    # ---------------------------
    # VALIDACIÓN
    # ---------------------------
    if df_final.empty:
        raise ValueError("El reporte de facturación quedó vacío. Verifica los datos.")

    # ---------------------------
    # LIMPIEZA GENERAL
    # ---------------------------
    df_final = df_final.replace({None: "", "nan": "", "None": ""})

    # ---------------------------
    # ORDEN DE COLUMNAS CORRECTO
    # ---------------------------
    columnas_ordenadas = [
        "ID CUENTA",
        "ID CLIENTE",
        "NOMBRE",
        "DOCUMENTO",
        "EMAIL",
        "TELÉFONO",
        "DIRECCIÓN",

        "ID CONTRATO PRINCIPAL",
        "CONTRATOS ASOCIADOS",
        "ESTADOS CONTRATO",
        "PLANES ASOCIADOS",

        "TOTAL FACTURAS",
        "FACTURAS PAGADAS",
        "MESES PAGADOS",
        "ÚLTIMO MES PAGADO",
        "VALOR TOTAL PAGADO",

        "FECHA EMISIÓN (FACTURA ÚLTIMA)",
        "PRIMER VENCIMIENTO",
        "SEGUNDO VENCIMIENTO",
        "PRÓXIMO CORTE",

        "ESTADO GENERAL",
    ]

    # Crear columnas faltantes (evita errores)
    for col in columnas_ordenadas:
        if col not in df_final.columns:
            df_final[col] = ""

    df_final = df_final[columnas_ordenadas].copy()

    # ---------------------------
    # TIPOS DE DATOS
    # ---------------------------
    columnas_fecha = [
        "FECHA EMISIÓN (FACTURA ÚLTIMA)",
        "PRIMER VENCIMIENTO",
        "SEGUNDO VENCIMIENTO",
        "PRÓXIMO CORTE",
    ]

    for col in columnas_fecha:
        if col in df_final.columns:
            df_final[col] = pd.to_datetime(df_final[col], errors="coerce")

    if "VALOR TOTAL PAGADO" in df_final.columns:
        df_final["VALOR TOTAL PAGADO"] = pd.to_numeric(
            df_final["VALOR TOTAL PAGADO"], errors="coerce"
        ).fillna(0)

    # ---------------------------
    # ✅ CORRECCIÓN #4 — ORDEN FINAL POR FECHA DE CORTE
    # Se ordena por PRÓXIMO CORTE ascendente para que los
    # vencimientos más próximos aparezcan primero (criterio
    # operativo ISP). Los registros sin fecha van al final.
    # Dentro del mismo corte, se ordena por ID CUENTA.
    # ---------------------------
    if "ID CUENTA" not in df_final.columns:
        df_final["ID CUENTA"] = ""
    if "PRÓXIMO CORTE" not in df_final.columns:
        df_final["PRÓXIMO CORTE"] = pd.NaT

    df_final = df_final.sort_values(
        by=["PRÓXIMO CORTE", "ID CUENTA"],
        ascending=[True, True],
        na_position="last",
        kind="mergesort"
    ).reset_index(drop=True)
    
    # ---------------------------
    # EXPORTACIÓN
    # ---------------------------
    fecha_salida = datetime.now().strftime("%Y-%m-%d")
    ruta_salida = BASE_SALIDA / f"reporte_facturacion_{fecha_salida}.xlsx"

    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        df_final.to_excel(writer, index=False, sheet_name="Facturacion_Clientes")

    _aplicar_formato_excel(ruta_salida)

    logger.info(f"Reporte generado correctamente: {ruta_salida.resolve()}")
    logger.info(f"Total clientes: {len(df_final)}")

    return ruta_salida

# ------------------------------------------------------------------
# EJECUCIÓN DIRECTA
# ------------------------------------------------------------------
if __name__ == "__main__":
    generar_reporte_facturacion()